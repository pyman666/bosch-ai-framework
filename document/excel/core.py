"""Excel 解析的"无业务"基础件: 读取/合并 cell/可见性过滤/泛用 unpivot 配置.

模块按职责分两块:
    - ``PyXL``: 一次性读 sheet 到 numpy.ndarray, 处理合并 cell, 提供 ``matrix_filled``.
    - ``PyExcel`` + ``ExcelConfig`` / ``FilterConfig`` / ``HeaderConfig``:
      用户可配置的行/列过滤 + 表头识别, 输出 ``pandas.DataFrame``.

业务专属的解析 (e.g. xpeng-zq 的 plan->rows) 不放这里, 放在 ``apdfi.excel.clients``.
"""
import json
from functools import cached_property
from typing import Type, TypeVar, Literal

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np
import pandas as pd

from pydantic import (
    BaseModel, Field, ValidationError, TypeAdapter
)
from fastapi import HTTPException


# ---------------------------------------------------------------------------
# config

class FilterBlock(BaseModel):
    refer_index: int = Field(..., ge=1)  # 1-based; converted to 0-based in FilterConfig.indices()
    refer_name: str

    def indices_mask(self, arr: np.ndarray) -> np.ndarray:
        return np.char.find(arr.astype(str), self.refer_name) >= 0


class FilterConfig(BaseModel):
    axis: Literal["row", "column"]
    mode: Literal["skip", "use"] = "skip"

    before: int = None
    after: int = None
    positions: list[int] = None  # 1-based; converted to 0-based in indices()
    blocks: list[FilterBlock] = None

    def indices(self, xl: "PyXL") -> np.ndarray:
        max_idx = xl.matrix_raw.shape[0 if self.axis == "row" else 1]
        mask = np.zeros(max_idx, dtype=bool)

        if self.before is not None:
            mask[:self.before] = True
        if self.after is not None:
            mask[self.after:] = True
        if self.positions:
            mask[[p - 1 for p in self.positions]] = True  # 1-based → 0-based

        if self.blocks:
            mat = xl.matrix_filled
            for block in self.blocks:
                idx = block.refer_index - 1  # 1-based → 0-based
                arr = mat[:, idx] if self.axis == "row" else mat[idx, :]
                mask |= block.indices_mask(arr)

        return np.where(mask)[0] if self.mode == "use" else np.where(~mask)[0]


class HeaderConfig(BaseModel):
    rows: int = 1
    row: int = 1
    block: dict[str, int] = None


class ExcelConfig(BaseModel):
    row_filter: FilterConfig = Field(
        default_factory=lambda: FilterConfig(axis="row"),
    )
    col_filter: FilterConfig = Field(
        default_factory=lambda: FilterConfig(axis="column"),
    )
    header_config: HeaderConfig = Field(
        default_factory=HeaderConfig,
    )


T = TypeVar("T", bound=BaseModel)


def json_config(data: str, model: Type[T]) -> T | None:
    """form 字段里塞 JSON 字符串 -> 对应的 pydantic 模型. 空字符串/None 返回 None."""
    if not data:
        return None

    try:
        return TypeAdapter(model).validate_json(data)
    except json.JSONDecodeError as e:
        raise HTTPException(400, f"配置项必须是合法 JSON: {e.msg}")
    except ValidationError as e:
        raise HTTPException(422, e.errors())


# ---------------------------------------------------------------------------
# raw sheet -> numpy

NA = frozenset({
    "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"
})


class PyXL:
    """一次性把 sheet 读成 ndarray, 缓存合并 cell 信息. 多个解析路径共享一个实例.

    隐藏行/列的判定在这里集中提供 (``hidden_rows_1b`` / ``hidden_cols_0b``), 上层
    (``PyExcel`` / 业务 pipeline / LLM skeleton) 全部复用这两个属性, 保持"默认跳隐藏"
    的颗粒度一致.
    """

    def __init__(
        self,
        file,
        sheet: int | str = 1,
    ):
        self.wb: Workbook = load_workbook(file, data_only=True)
        self.ws: Worksheet = (
            self.wb[sheet]
            if isinstance(sheet, str)
            else self.wb.worksheets[sheet - 1]
        )

        self.matrix_raw: np.ndarray = np.array(
            list(self.ws.iter_rows(
                min_row=1,
                max_row=self.ws.max_row,
                min_col=1,
                max_col=self.ws.max_column,
                values_only=True,
            )),
            dtype=object,
        )

        ranges = list(self.ws.merged_cells.ranges)
        self.merged_bounds = (
            np.array([rng.bounds for rng in ranges], dtype=int)
            if ranges
            else np.empty((0, 4), dtype=int)
        )

    @cached_property
    def matrix_filled(self) -> np.ndarray:
        """合并 cell 区域的值填到每个子格上, 拿来做行/列过滤更直观."""
        mat = self.matrix_raw.copy()
        if self.merged_bounds.size == 0:
            return mat

        min_row = self.merged_bounds[:, 1] - 1
        max_row = self.merged_bounds[:, 3]
        min_col = self.merged_bounds[:, 0] - 1
        max_col = self.merged_bounds[:, 2]

        for i in range(len(self.merged_bounds)):
            val = mat[min_row[i], min_col[i]]
            if val is not None:
                mat[min_row[i]: max_row[i], min_col[i]: max_col[i]] = val
        return mat

    @cached_property
    def hidden_rows_1b(self) -> set[int]:
        """openpyxl 的 row_dimensions 是按行号 (1-based) 索引的, 直接转 set."""
        return {
            i for i, dim in self.ws.row_dimensions.items()
            if dim.hidden or (dim.height is not None and dim.height <= 0)
        }

    @cached_property
    def hidden_cols_0b(self) -> set[int]:
        """按 docs 旧逻辑逐列判断可见性, 保留列维度继承行为."""
        ws = self.ws
        max_col = ws.max_column
        default_width = ws.sheet_format.defaultColWidth

        visible = np.ones(max_col, dtype=bool)
        width = default_width
        hidden = False

        from openpyxl.utils import get_column_letter

        for i in range(1, max_col + 1):
            letter = get_column_letter(i)
            dim = ws.column_dimensions.get(letter)

            if dim is not None:
                if dim.width is not None:
                    width = dim.width
                if dim.hidden is not None:
                    hidden = dim.hidden

            if hidden or (width is not None and width <= 0):
                visible[i - 1] = False

        return set(np.where(~visible)[0].tolist())


# ---------------------------------------------------------------------------
# filtered sheet -> pandas

class PyExcel:
    """配合 ``ExcelConfig`` 把 ``PyXL`` 的矩阵过滤 + 提取表头, 出一个 DataFrame."""

    def __init__(
        self,
        excel: PyXL,
        read_hidden: bool = False,
        config: ExcelConfig = None,
    ):
        self.xl = excel
        self.read_hidden: bool = read_hidden
        self.config = config

    @cached_property
    def matrix(self) -> np.ndarray:
        if self.config is None:
            return np.empty((0, 0))

        mat = self.xl.matrix_filled.copy()

        row_idx = self.config.row_filter.indices(self.xl)
        col_idx = self.config.col_filter.indices(self.xl)

        if not self.read_hidden:
            # 复用 PyXL 上集中维护的 hidden 集合, 颗粒度跟业务侧 pipeline 一致.
            row_hidden = np.array([i - 1 for i in self.xl.hidden_rows_1b], dtype=int)
            col_hidden = np.array(sorted(self.xl.hidden_cols_0b), dtype=int)

            row_idx = np.setdiff1d(row_idx, row_hidden)
            col_idx = np.setdiff1d(col_idx, col_hidden)

        mat = mat[np.ix_(row_idx, col_idx)]
        mat[np.isin(mat, list(NA))] = np.nan
        return mat

    def to_frame(self) -> pd.DataFrame:
        if self.config is None:
            return pd.DataFrame()

        cfg = self.config.header_config
        df = pd.DataFrame(self.matrix[cfg.rows:])

        if not cfg.block:
            df.columns = self.matrix[cfg.row - 1].astype(str)
        else:
            names, rows = zip(*cfg.block.items())
            df.columns = pd.MultiIndex.from_arrays(
                self.matrix[[i - 1 for i in rows]],
                names=names,
            )
            df.columns = df.columns.map(lambda x: tuple(map(str, x)))
        return df
